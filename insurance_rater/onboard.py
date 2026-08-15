"""Adapter proposer: the LLM reads a rate-card workbook and *proposes* how to
resolve rates from it -- it never resolves anything itself.

    python -m insurance_rater.onboard <workbook.xlsx>

Pipeline: dump every sheet to text -> one JSON-mode LLM call (same free-model
failover chain as extraction) -> schema/coordinate validation against the real
workbook (hallucinated sheets or cells are reported, not trusted) -> cross-check
against what the hand-written resolver for that insurer actually uses -> write
<workbook dir>/adapter.proposed.json for human review.

Nothing in the rating path reads the proposal. The trust chain is:
LLM proposes -> validation + human review verify -> (future) code executes.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import warnings

import openpyxl
from openpyxl.utils import get_column_letter

from . import insurers, llm

_MAX_COLS = 50
_HEAD_ROWS = 45
_TAIL_ROWS = 8
_CELL_CHARS = 40

_SYSTEM = """You are onboarding an Indian motor-insurance broker commission rate card \
(an Excel workbook) so that deterministic code can later resolve rates from it. \
Read the sheet dumps and return ONLY a JSON object:

{
 "insurer": "<insurer name as written in the workbook, or null>",
 "sheets": [{"name": "<sheet>", "purpose": "<one line: what this sheet holds>"}],
 "rate_sheets": [{
   "sheet": "<sheet holding rate cells>",
   "product": "<vehicle product it rates, e.g. 'private car', 'two-wheeler', 'school bus'>",
   "header_row": <1-based row number of the column-header row>,
   "dimensions": [{
     "name": "<axis a lookup must pin, e.g. 'fuel', 'business type', 'region'>",
     "values": ["<the values as written in the sheet>"],
     "derive_from_policy": "<how a policy schedule would pin this, one line>"}],
   "notes": "<units, payout basis, anything a resolver must know>"}],
 "rto_mappings": [{"sheet": "<sheet>", "maps": "<what -> what, e.g. 'RTO code -> region city'>"}],
 "footnotes": [{"sheet": "<sheet>", "cell": "<A1-style cell>",
                "meaning": "<the rate-modifying rule in that cell, one line>"}]
}

Rules: report only what is present in the dumps; cite real sheet names and cells. \
Rates in the sheets may be fractions (0.25) or percents (25) -- note which in "notes". \
List every sheet, but only genuine rate grids belong in "rate_sheets". A footnote \
belongs in "footnotes" only if it changes a rate (deductions, caps, payout basis)."""


def dump_sheet(ws) -> str:
    ncols = min(ws.max_column, _MAX_COLS)
    lines = [f"### Sheet: {ws.title} ({ws.max_row} rows x {ws.max_column} cols)"]
    rows = range(1, ws.max_row + 1)
    if ws.max_row > _HEAD_ROWS + _TAIL_ROWS:
        rows = [*range(1, _HEAD_ROWS + 1),
                *range(ws.max_row - _TAIL_ROWS + 1, ws.max_row + 1)]
    prev = None
    for r in rows:
        if prev is not None and r != prev + 1:
            lines.append(f"... rows {prev + 1}-{r - 1} omitted ...")
        prev = r
        # coordinate=value pairs so the model can cite real cells
        vals = [f"{get_column_letter(c)}{r}={str(ws.cell(r, c).value)[:_CELL_CHARS]}"
                for c in range(1, ncols + 1) if ws.cell(r, c).value is not None]
        if vals:
            lines.append(" | ".join(vals))
    return "\n".join(lines)


def propose(path: str) -> dict:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)
    dump = "\n\n".join(dump_sheet(ws) for ws in wb)
    raw = llm._post({
        "temperature": 0,
        "reasoning_effort": "low",
        "response_format": {"type": "json_object"},
        # Go Digit's workbook has ~20 sheets; 4000 truncated the JSON mid-string
        "max_tokens": 12000,
        "messages": [{"role": "system", "content": _SYSTEM},
                     {"role": "user", "content": f"Workbook file: {os.path.basename(path)}\n\n{dump}"}],
    })
    raw["_meta"] = {"workbook": os.path.basename(path), "model": llm._last_model}
    return raw


_CELL_RE = re.compile(r"^[A-Z]{1,3}[0-9]{1,7}$")


def validate(proposal: dict, wb) -> list[str]:
    """Reject hallucinations: every sheet and cell the LLM cites must exist.

    Sheet names match whitespace-insensitively: real workbooks have names like
    'School & Staff Bus ' (trailing space) that models normalize away.
    """
    issues = []
    canon = {" ".join(n.split()): n for n in wb.sheetnames}

    def real(name) -> str | None:
        return canon.get(" ".join(str(name or "").split()))

    for s in proposal.get("sheets", []):
        if real(s.get("name")) is None:
            issues.append(f"sheets: '{s.get('name')}' is not in the workbook")
    for rs in proposal.get("rate_sheets", []):
        sheet = real(rs.get("sheet"))
        if sheet is None:
            issues.append(f"rate_sheets: sheet '{rs.get('sheet')}' is not in the workbook")
            continue
        hdr = rs.get("header_row")
        if not isinstance(hdr, int) or not 1 <= hdr <= wb[sheet].max_row:
            issues.append(f"rate_sheets: '{sheet}' header_row {hdr!r} out of range")
    for m in proposal.get("rto_mappings", []):
        if real(m.get("sheet")) is None:
            issues.append(f"rto_mappings: sheet '{m.get('sheet')}' is not in the workbook")
    for fn in proposal.get("footnotes", []):
        sheet, cell = real(fn.get("sheet")), str(fn.get("cell") or "")
        if sheet is None:
            issues.append(f"footnotes: sheet '{fn.get('sheet')}' is not in the workbook")
        elif not _CELL_RE.match(cell):
            issues.append(f"footnotes: '{cell}' is not an A1-style cell")
        elif wb[sheet][cell].value is None:
            issues.append(f"footnotes: {sheet}!{cell} is empty in the workbook")
    return issues


# What the hand-written resolvers actually read, per insurer -- the ground truth
# the proposal is compared against. (HDFC's card is a PDF; xlsx-only tool.)
_KNOWN = {
    "reliance": {"rate_sheets": {"PRIVATE CAR COMP, SAOD & STP"}, "rto": {"RTO List"}},
    "tata_aig": {"rate_sheets": {"Pvtcar"}, "rto": set(),
                 "dimension_values": {"Brand New", "Renewal", "Rollover"}},
    "go_digit": {"rate_sheets": {"4W SATP"}, "rto": {"4W  RTO"}},
}


def cross_check(proposal: dict, insurer_key: str | None) -> list[str]:
    known = _KNOWN.get(insurer_key or "")
    if not known:
        return [f"no hand-written resolver to compare against (insurer={insurer_key})"]
    out = []
    proposed_rates = {rs.get("sheet") for rs in proposal.get("rate_sheets", [])}
    for sheet in known["rate_sheets"]:
        hit = "AGREES" if sheet in proposed_rates else "MISSED"
        out.append(f"{hit}: resolver rate sheet '{sheet}' "
                   f"{'found' if hit == 'AGREES' else 'not proposed'}")
    proposed_rto = {m.get("sheet") for m in proposal.get("rto_mappings", [])}
    for sheet in known["rto"]:
        hit = "AGREES" if sheet in proposed_rto else "MISSED"
        out.append(f"{hit}: resolver RTO sheet '{sheet}'")
    want = known.get("dimension_values")
    if want:
        got = {v for rs in proposal.get("rate_sheets", [])
               for d in rs.get("dimensions", []) for v in d.get("values", [])}
        hit = "AGREES" if want <= got else "MISSED"
        out.append(f"{hit}: business-type dimension values {sorted(want)}")
    extra = proposed_rates - known["rate_sheets"]
    if extra:
        out.append(f"EXTRA: proposed rate sheets beyond the resolver's scope: {sorted(extra)} "
                   "(expected -- resolvers only cover private car)")
    return out


def _detect_insurer(path: str, wb) -> str | None:
    head = path + " " + " ".join(wb.sheetnames)
    for ins in insurers.REGISTRY:
        if any(re.search(fp, head, re.I) for fp in ins.fingerprints):
            return ins.key
    return None


def main(argv=None):
    ap = argparse.ArgumentParser(prog="insurance_rater.onboard")
    ap.add_argument("workbook", help="rate-card .xlsx to propose an adapter for")
    args = ap.parse_args(argv)

    proposal = propose(args.workbook)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(args.workbook, data_only=True)
    issues = validate(proposal, wb)
    checks = cross_check(proposal, _detect_insurer(args.workbook, wb))

    out_path = os.path.join(os.path.dirname(args.workbook) or ".", "adapter.proposed.json")
    proposal["_meta"]["validation_issues"] = issues
    proposal["_meta"]["cross_check"] = checks
    with open(out_path, "w") as f:
        json.dump(proposal, f, indent=2, ensure_ascii=False)

    print(f"proposal written to {out_path}  (model: {proposal['_meta']['model']})")
    print(f"\nvalidation: {len(issues)} issue(s)")
    for i in issues:
        print(f"  ! {i}")
    print("\ncross-check vs hand-written resolver:")
    for c in checks:
        print(f"  {c}")
    if issues:
        sys.exit(1)


if __name__ == "__main__":
    main()
