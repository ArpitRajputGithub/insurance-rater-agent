"""Shared helpers for the per-insurer grid resolvers.

Rate-card file locations, workbook loading, citation builders, and the two
component constructors (`_na` / `_unsupported`) every resolver returns. Nothing
here is specific to one insurer's card -- that logic lives in the sibling
modules (go_digit, hdfc, reliance, tata).
"""
from __future__ import annotations

import os
import warnings

import openpyxl
from openpyxl.utils import get_column_letter

from .. import insurers
from ..models import Citation, ComponentRate, Status


def grid_path(insurer_key: str, raters_dir: str) -> str:
    return os.path.join(raters_dir, insurers.grid_file(insurer_key))


def _source(insurer_key: str) -> str:  # basename doubles as the citation source
    return os.path.basename(insurers.grid_file(insurer_key))


def _wb(insurer_key: str, raters_dir: str):
    # Suppress openpyxl's warning about an embedded image; it does not affect cell values.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(grid_path(insurer_key, raters_dir), data_only=True)


def _xls_cite(insurer_key, sheet, row, col, detail):
    loc = f"{sheet}!{get_column_letter(col)}{row}"
    return Citation(source=_source(insurer_key), locator=loc, detail=detail)


# --- label-anchored lookups -------------------------------------------------
# Rate cards are re-issued monthly. Numbers change harmlessly (we read cells,
# not constants), but a column can move or a header can be renamed. Anchoring on
# header *text* instead of a fixed index survives a reshuffle; when an anchor is
# gone we refuse loudly (see _struct_changed) rather than read a neighbour and
# return a plausible-but-wrong rate.
def _norm(s) -> str:
    return " ".join(str(s if s is not None else "").split()).strip().lower()


def find_row(ws, *labels, limit: int = 40):
    """First row index (1-based) whose cells contain every label, or None.

    Each label matches a cell that equals it or contains it (both normalised),
    so 'SA OD' finds a 'SA OD' header regardless of surrounding whitespace/case.
    """
    want = [_norm(x) for x in labels]
    for r in range(1, min(ws.max_row, limit) + 1):
        cells = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if all(any(w == c or w in c for c in cells) for w in want):
            return r
    return None


def header_map(ws, row: int) -> dict:
    """normalised header text -> column index, for the given header row."""
    return {_norm(ws.cell(row, c).value): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row, c).value not in (None, "")}


def _struct_changed(component: str, insurer_name: str, what: str) -> ComponentRate:
    return _unsupported(component, f"{insurer_name} rate-card structure changed: could "
                        f"not locate {what}. Refusing to guess a rate.")


def _na(component: str, reason: str) -> ComponentRate:
    """Component that does not exist for this policy (never 0%)."""
    return ComponentRate(component=component, applicable=False,
                         status=Status.RESOLVED, rate_percent=None, reason=reason)


def _unsupported(component: str, reason: str, citations=None) -> ComponentRate:
    return ComponentRate(component=component, applicable=True,
                         status=Status.UNSUPPORTED, rate_percent=None,
                         reason=reason, citations=citations or [])
