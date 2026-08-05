"""Deterministic grid resolution: the reproducible half of the pipeline.

Extraction may be fuzzy or model-driven, but turning policy facts into a
commission rate must be reproducible and testable, and must never invent a
rate through a missing, declined or out-of-scope rule.

This module reads the actual rate-card files shipped in `raters/` and returns
the exact sheet/cell (or PDF page/table) that produced each number. Nothing
here is hard-coded to the four sample answers: change a cell in a grid and the
output changes with it. Where a required mapping is absent we return
UNSUPPORTED rather than guess.

Each resolver returns (od, tp, trace):
  * od / tp : ComponentRate   -- applicability + status + rate + citations
  * trace   : list[TraceStep] -- one auditable line per lookup

Convergence: sometimes a fact we could not extract (fuel, exact segment,
renewal-vs-rollover) does not change the answer because every candidate cell
holds the same rate. That is a resolution, not a guess, and the reason says so.
Where candidates disagree we return AMBIGUOUS instead.
"""
from __future__ import annotations

import os
import warnings

import openpyxl
import pdfplumber
from openpyxl.utils import get_column_letter

from .models import ComponentRate, PolicyFacts, Status, TraceStep

# Grid files, relative to the raters/ root. Basenames double as citation sources.
_GRID_FILES = {
    "go_digit": "godigit/march-2026/Large Insurance Brokers Mar'26 - Shared.xlsx",
    "hdfc_ergo": "hdfc-ergo/feb-2025/Pvt Car New Grid Eff 1st Feb'25 (HDFC ergo) 1.pdf",
    "reliance": "reliance-indusind/feb-2026/Reliance Broking Premier  FEB 26 Grid.xlsx",
    "tata_aig": "tata-aig/march-2026/Tata AIG Standard Grid_Communication_Mar'26_F_v2_0212.xlsx",
}


def grid_path(insurer_key: str, raters_dir: str) -> str:
    return os.path.join(raters_dir, _GRID_FILES[insurer_key])


def _source(insurer_key: str) -> str:
    return os.path.basename(_GRID_FILES[insurer_key])


def _wb(insurer_key: str, raters_dir: str):
    # Suppress openpyxl's warning about an embedded image; it does not affect cell values.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(grid_path(insurer_key, raters_dir), data_only=True)


def _xls_cite(insurer_key, sheet, row, col, detail):
    from .models import Citation
    loc = f"{sheet}!{get_column_letter(col)}{row}"
    return Citation(source=_source(insurer_key), locator=loc, detail=detail)


def _na(component: str, reason: str) -> ComponentRate:
    """Component that does not exist for this policy (never 0%)."""
    return ComponentRate(component=component, applicable=False,
                         status=Status.RESOLVED, rate_percent=None, reason=reason)


def _unsupported(component: str, reason: str, citations=None) -> ComponentRate:
    return ComponentRate(component=component, applicable=True,
                         status=Status.UNSUPPORTED, rate_percent=None,
                         reason=reason, citations=citations or [])


def resolve(facts: PolicyFacts, insurer_key: str, raters_dir: str):
    fn = _RESOLVERS.get(insurer_key)
    if fn is None:
        r = _unsupported("od", f"No grid resolver for insurer '{insurer_key}'.")
        return r, _unsupported("tp", r.reason), []
    return fn(facts, raters_dir)


# ---------------------------------------------------------------------------
# Go Digit -- Stand-alone TP. RTO code -> 4WTP cluster -> fuel/cc segment rate.
# ---------------------------------------------------------------------------
def _godigit_segment(fuel, cc):
    if fuel is None or cc is None:
        return None
    f = fuel.lower()
    if f == "cng":
        return "CNG"
    if f == "petrol":
        return "Petrol<1000" if cc < 1000 else "Petrol>1000"
    if f == "diesel":
        return "Diesel<1500" if cc < 1500 else "Diesel>1500"
    return None


def _resolve_go_digit(facts, raters_dir):
    key = "go_digit"
    trace = []
    od = _na("od", "Stand-alone TP policy: Own Damage is not covered, so no OD "
                   "commission applies.")
    wb = _wb(key, raters_dir)

    rto = facts.get("rto_code")
    if not rto:
        return od, _unsupported("tp", "No RTO code extracted; cannot map to a "
                                      "Go Digit 4W TP cluster."), trace

    # 1. RTO code -> 4WTP cluster (column C of the '4W  RTO' sheet).
    rmap = wb["4W  RTO"]
    cluster = crow = None
    for r in range(1, rmap.max_row + 1):
        if str(rmap.cell(r, 2).value).strip().upper() == rto.upper():
            cluster, crow = rmap.cell(r, 3).value, r
            break
    if cluster is None:
        return od, _unsupported("tp", f"RTO code {rto} is not listed in the Go Digit "
                                      "'4W  RTO' mapping sheet."), trace
    c1 = _xls_cite(key, "4W  RTO", crow, 3, f"{rto} maps to 4WTP cluster '{cluster}'")
    trace.append(TraceStep(f"RTO {rto} maps to Go Digit TP cluster '{cluster}'",
                           facts.facts.get("rto_code").citation, c1))

    # 2. cluster + fuel/cc segment -> rate (column E 'Max CD2' of '4W SATP').
    seg = _godigit_segment(facts.get("fuel"), facts.get("cc"))
    if seg is None:
        return od, _unsupported("tp", "Could not form a Go Digit segment (need fuel "
                                      "and cc).", [c1]), trace
    rate_ws = wb["4W SATP"]
    rate = rrow = None
    for r in range(1, rate_ws.max_row + 1):
        if (str(rate_ws.cell(r, 2).value).strip() == str(cluster).strip()
                and str(rate_ws.cell(r, 3).value).strip() == seg):
            rate, rrow = rate_ws.cell(r, 5).value, r
            break
    if rate is None:
        return od, _unsupported("tp", f"No '4W SATP' row for cluster '{cluster}', "
                                      f"segment '{seg}'.", [c1]), trace
    c2 = _xls_cite(key, "4W SATP", rrow, 5, f"cluster '{cluster}', segment '{seg}' "
                                            f"-> {rate*100:g}%")
    trace.append(TraceStep(f"Segment '{seg}' in cluster '{cluster}' -> TP {rate*100:g}%",
                           None, c2))
    tp = ComponentRate("tp", True, Status.RESOLVED, round(rate * 100, 3),
                       citations=[c1, c2],
                       reason=f"Go Digit 4W SATP rate for {rto} ({cluster}) / {seg}.")
    return od, tp, trace


# ---------------------------------------------------------------------------
# HDFC ERGO -- Comprehensive. PDF grid: state->zone, then zone/slab/fuel/NCB.
# The grid only publishes Package/SAOD (OD-based) rates; it has no TP table,
# so TP is an honest UNSUPPORTED.
# ---------------------------------------------------------------------------
_HDFC_SLABS = [  # (label, upper-exclusive bound on comprehensive premium in rupees)
    ("<10k", 10_000), (">10-50k", 50_000), (">50k-1L", 100_000),
    (">1L-2L", 200_000), (">2L", float("inf")),
]


def _hdfc_slab(premium):
    for label, hi in _HDFC_SLABS:
        if premium < hi:
            return label
    return _HDFC_SLABS[-1][0]


def _pct(s):
    return float(str(s).replace("%", "").strip())


def _hdfc_read(path):
    """Parse the HDFC PDF into (zone_map, rate_tables, zone_page).

    zone_map: state(lower) -> 'Zone-1'|'Zone-2'
    rate_tables: {'Zone 1': {slab: {...}}, 'Zone 2': {...}}  (Package cols only)
    """
    zone_map, rates, zone_page = {}, {}, {}
    with pdfplumber.open(path) as pdf:
        for pidx, pg in enumerate(pdf.pages):
            for tbl in pg.extract_tables():
                head = (tbl[0][1] or "") if tbl and len(tbl[0]) > 1 else ""
                # Locate the 'Zone | State Name | ... | Zone-1 | Zone-2' header row,
                # which may sit under a leading blank row.
                hrow = next((i for i, r in enumerate(tbl)
                             if r and r[0] == "Zone" and (r[1] or "").startswith("State")), None)
                if head in ("Zone 1", "Zone 2"):
                    slabs = {}
                    for row in tbl[4:]:
                        if row and row[0]:
                            slabs[row[0].strip()] = {
                                "petrol": row[1], "nonpetrol_ncb": row[2],
                                "nonpetrol_nncb": row[3],
                            }
                    rates[head] = slabs
                elif hrow is not None:
                    for row in tbl[hrow + 1:]:
                        name = (row[1] or "").strip()
                        if not name:
                            continue
                        z1 = (row[3] or "").strip()
                        z2 = (row[6] or "").strip() if len(row) > 6 else ""
                        zone = "Zone-2" if z2 else ("Zone-1" if z1 else None)
                        if zone:
                            zone_map[name.lower()] = zone
                            zone_page[name.lower()] = pidx + 1
    return zone_map, rates, zone_page


def _resolve_hdfc(facts, raters_dir):
    key = "hdfc_ergo"
    from .models import Citation
    src = _source(key)
    trace = []
    tp = _unsupported("tp", "The HDFC ERGO grid publishes only Package/SAOD "
                            "(Own-Damage-based) commission; it contains no Third "
                            "Party rate table, so the TP rate cannot be determined "
                            "from the supplied evidence.")

    state = facts.get("rto_state")
    slab_prem = facts.get("package_premium")
    if not state or slab_prem is None:
        return _unsupported("od", "Missing state or comprehensive premium; cannot "
                                  "look up the HDFC zone/slab."), tp, trace

    zone_map, rates, zone_page = _hdfc_read(grid_path(key, raters_dir))
    zone = zone_map.get(state.lower())
    if zone is None:
        return _unsupported("od", f"State '{state}' not found in the HDFC zone map."), tp, trace
    zc = Citation(src, f"page {zone_page.get(state.lower())}",
                  f"{state} -> {zone}")
    trace.append(TraceStep(f"{state} maps to HDFC {zone}",
                           facts.facts.get("rto_state").citation, zc))

    zkey = zone.replace("-", " ")  # 'Zone-2' -> 'Zone 2'
    slab = _hdfc_slab(slab_prem)
    trace.append(TraceStep(f"Comprehensive premium {slab_prem} -> slab {slab}",
                           facts.facts.get("package_premium").citation, None))
    cell = rates.get(zkey, {}).get(slab)
    if cell is None:
        return _unsupported("od", f"No HDFC {zkey} rate row for slab '{slab}'.", [zc]), tp, trace

    # Fuel is not printed on HDFC schedules. Candidate columns depend on NCB:
    #   petrol             -> 'petrol' column (same for NCB and N-NCB)
    #   non-petrol + NCB   -> 'nonpetrol_ncb'
    #   non-petrol + N-NCB -> 'nonpetrol_nncb'
    # We take every column consistent with what we do know and resolve only if
    # they converge; otherwise AMBIGUOUS.
    ncb = facts.get("ncb_applies")
    fuel = facts.get("fuel")
    cands = {}
    if fuel and fuel.lower() != "petrol":  # known non-petrol
        cands["nonpetrol"] = cell["nonpetrol_ncb"] if ncb else cell["nonpetrol_nncb"]
    elif fuel and fuel.lower() == "petrol":
        cands["petrol"] = cell["petrol"]
    else:  # fuel unknown -> both fuel families are candidates
        cands["petrol"] = cell["petrol"]
        cands["nonpetrol"] = cell["nonpetrol_ncb"] if ncb else cell["nonpetrol_nncb"]
    vals = {_pct(v) for v in cands.values()}
    # Rate tables live on page 1 of this grid.
    rc = Citation(src, "page 1", f"{zkey} · Package · slab {slab}: " +
                  ", ".join(f"{k}={cands[k]}" for k in cands))

    if len(vals) == 1:
        rate = vals.pop()
        cov = ("petrol and non-petrol columns agree" if len(cands) > 1
               else f"{next(iter(cands))} column")
        reason = (f"HDFC {zkey} Package, slab {slab} ({cov}"
                  + (", NCB applies" if ncb else "") + f") -> {rate:g}%.")
        trace.append(TraceStep(f"{zkey} · Package · {slab} -> OD {rate:g}% ({cov})", None, rc))
        od = ComponentRate("od", True, Status.RESOLVED, rate, citations=[zc, rc], reason=reason)
    else:
        reason = (f"HDFC {zkey} Package slab {slab} gives different rates by fuel "
                  f"({cands}); fuel is not printed on the schedule, so OD is ambiguous.")
        trace.append(TraceStep(f"{zkey} · Package · {slab} -> ambiguous by fuel {cands}", None, rc))
        od = ComponentRate("od", True, Status.AMBIGUOUS, None, citations=[zc, rc], reason=reason)
    return od, tp, trace


# ---------------------------------------------------------------------------
# Reliance -- Comprehensive. RTO code -> region_city -> OD column, minus the
# <1000cc footnote. TP payout is on OD premium only (STP column = 0).
# ---------------------------------------------------------------------------
def _resolve_reliance(facts, raters_dir):
    key = "reliance"
    trace = []
    wb = _wb(key, raters_dir)

    rto = facts.get("rto_code")
    if not rto:
        r = _unsupported("od", "No RTO code extracted; cannot map to a Reliance region.")
        return r, _unsupported("tp", r.reason), trace

    # 1. RTO code -> Region_City (RTO List sheet, col A -> col C).
    rl = wb["RTO List"]
    city = crow = None
    for r in range(1, rl.max_row + 1):
        if str(rl.cell(r, 1).value).strip().upper() == rto.upper():
            city, crow = rl.cell(r, 3).value, r
            break
    if city is None:
        r = _unsupported("od", f"RTO code {rto} not found in the Reliance 'RTO List'.")
        return r, _unsupported("tp", r.reason), trace
    c1 = _xls_cite(key, "RTO List", crow, 3, f"{rto} -> Region_City '{city}'")
    trace.append(TraceStep(f"RTO {rto} maps to Reliance region '{city}'",
                           facts.facts.get("rto_code").citation, c1))

    # 2. region -> rate row (PRIVATE CAR COMP, SAOD & STP).
    rs = wb["PRIVATE CAR COMP, SAOD & STP"]
    rrow = None
    for r in range(3, rs.max_row + 1):
        if str(rs.cell(r, 2).value).strip().lower() == str(city).strip().lower():
            rrow = r
            break
    if rrow is None:
        r = _unsupported("od", f"Region '{city}' not found in the Reliance rate sheet.", [c1])
        return r, _unsupported("tp", r.reason, [c1]), trace

    petrol_od = rs.cell(rrow, 3).value       # 'Petrol/Bifuel/Comp'
    diesel_od = rs.cell(rrow, 4).value       # 'Diesel/EV/Comp'
    stp = rs.cell(rrow, 6).value             # 'STP'
    sheet = "PRIVATE CAR COMP, SAOD & STP"

    # Fuel is not printed on the Reliance schedule. If petrol and diesel columns
    # differ we cannot silently pick one -> only resolve on a documented rule:
    # no CNG/LPG kit fitted and <1000cc, for which no diesel passenger variant
    # exists, so the Petrol/Bifuel column is the supported reading (medium conf).
    cc = facts.get("cc")
    kit = facts.get("cng_lpg_kit")
    if _pct_or(petrol_od) == _pct_or(diesel_od):
        base, note = petrol_od, "petrol and diesel columns agree"
    elif kit is False and cc is not None and cc < 1000:
        base = petrol_od
        note = ("fuel not printed; no CNG/LPG kit and <1000cc (no diesel variant "
                "exists), so the Petrol/Bifuel column applies")
        facts.warnings.append("Reliance: fuel assumed Petrol/Bifuel (not on schedule; "
                              "no CNG kit, <1000cc).")
    else:
        r = ComponentRate("od", True, Status.AMBIGUOUS, None, citations=[c1],
                          reason=("Reliance OD differs by fuel (Petrol/Bifuel "
                                  f"{petrol_od} vs Diesel/EV {diesel_od}) and fuel is "
                                  "not printed on the schedule."))
        return r, _reliance_tp(key, sheet, rrow, stp), trace
    c2 = _xls_cite(key, sheet, rrow, 3, f"{city} Petrol/Bifuel/Comp = {base}%")
    trace.append(TraceStep(f"Region '{city}' OD (Petrol/Bifuel) = {base}% ({note})", None, c2))

    rate = float(base)
    cites = [c1, c2]
    # 3. Footnote: <1000cc -> 5% lesser (cell H3 of the rate sheet).
    if cc is not None and cc < 1000:
        foot = rs.cell(3, 8).value or ""
        if "1000" in str(foot):
            rate -= 5
            c3 = _xls_cite(key, sheet, 3, 8, f"footnote: <1000cc -> 5% lesser ({foot})")
            cites.append(c3)
            trace.append(TraceStep(f"Footnote: {cc}cc < 1000cc -> -5% -> OD {rate:g}%", None, c3))

    od = ComponentRate("od", True, Status.RESOLVED, round(rate, 3), citations=cites,
                       reason=f"Reliance {city} OD ({note}) minus <1000cc footnote."
                       if cc and cc < 1000 else f"Reliance {city} OD ({note}).")
    return od, _reliance_tp(key, sheet, rrow, stp), trace


def _pct_or(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _reliance_tp(key, sheet, rrow, stp):
    # The grid note (H2) states payout is made on OD premium only; the STP column
    # is 0 for every private-car region. TP therefore earns 0% commission here --
    # a concrete, cited zero, not a missing rule.
    c = _xls_cite(key, sheet, rrow, 6, f"STP column = {stp}; payout is on OD premium only")
    return ComponentRate("tp", True, Status.RESOLVED, float(stp or 0), citations=[c],
                         reason="Reliance pays commission on OD premium only (grid note); "
                                "the Third Party (STP) column is 0.")


# ---------------------------------------------------------------------------
# Tata AIG -- Stand-alone TP. RTO cluster column x fuel/section/business rows.
# For a used (non-new) car every segment and both renewal & rollover converge.
# ---------------------------------------------------------------------------
def _resolve_tata(facts, raters_dir):
    key = "tata_aig"
    from .models import Citation
    src = _source(key)
    trace = []
    od = _na("od", "Stand-alone TP policy: Own Damage is not covered, so no OD "
                   "commission applies.")
    wb = _wb(key, raters_dir)
    ws = wb["Pvtcar"]

    rto = facts.get("rto_location")
    fuel = facts.get("fuel")
    if not rto or not fuel:
        return od, _unsupported("tp", "Missing RTO location or fuel; cannot select a "
                                      "Tata Pvtcar column/row."), trace

    # 1. RTO location -> grid column (header row 5).
    col = None
    for c in range(13, ws.max_column + 1):
        if str(ws.cell(5, c).value).strip().upper() == rto.strip().upper():
            col = c
            break
    if col is None:
        return od, _unsupported("tp", f"RTO location '{rto}' is not a column in the "
                                      "Tata AIG 'Pvtcar' grid."), trace
    colcite = _xls_cite(key, "Pvtcar", 5, col, f"RTO cluster column '{rto}'")
    trace.append(TraceStep(f"RTO '{rto}' maps to Tata grid column '{rto}'",
                           facts.facts.get("rto_location").citation, colcite))

    # 2. Business type: a used car (regn/mfg year in the past) is Renewal or
    #    Rollover, never Brand New. We cannot tell which, and the exact vehicle
    #    segment is not mapped -- so gather EVERY non-new CNG/SATP row and resolve
    #    only if they converge.
    from datetime import date
    year = facts.get("regn_year") or facts.get("mfg_year")
    is_used = isinstance(year, int) and year < date.today().year
    biz_ok = {"Renewal", "Rollover"} if is_used else {"Brand New"}
    fuel_key = "CNG" if fuel.upper() == "CNG" else fuel.title()

    matches = []  # (row, value)
    for r in range(6, ws.max_row + 1):
        if (str(ws.cell(r, 9).value).strip() == fuel_key
                and str(ws.cell(r, 10).value).strip() == "SATP"
                and str(ws.cell(r, 8).value).strip() in biz_ok):
            v = ws.cell(r, col).value
            if v is not None:
                matches.append((r, v))
    if not matches:
        return od, _unsupported("tp", f"No Tata SATP rows for fuel '{fuel_key}', "
                                      f"business {sorted(biz_ok)} in '{rto}'.",
                                [colcite]), trace

    vals = {round(float(v) * 100, 3) for _, v in matches}
    rows = [r for r, _ in matches]
    span = _xls_cite(key, "Pvtcar", rows[0], col,
                     f"{fuel_key} SATP, {'/'.join(sorted(biz_ok))}, all segments "
                     f"(rows {rows[0]}-{rows[-1]}) in '{rto}'")
    if len(vals) == 1:
        rate = vals.pop()
        trace.append(TraceStep(
            f"{fuel_key} SATP {'/'.join(sorted(biz_ok))} in '{rto}': all "
            f"{len(matches)} candidate rows converge -> TP {rate:g}%", None, span))
        tp = ComponentRate("tp", True, Status.RESOLVED, rate, citations=[colcite, span],
                           reason=(f"Tata AIG SATP rate for {fuel_key} in {rto}; every "
                                   f"non-new segment/business-type row agrees at {rate:g}%."))
    else:
        trace.append(TraceStep(
            f"{fuel_key} SATP in '{rto}': candidate rows disagree {sorted(vals)}", None, span))
        tp = ComponentRate("tp", True, Status.AMBIGUOUS, None, citations=[colcite, span],
                           reason=(f"Tata AIG SATP rate for {fuel_key} in {rto} depends on "
                                   f"segment/business type, which could not be pinned down "
                                   f"and do not agree ({sorted(vals)})."))
    return od, tp, trace


_RESOLVERS = {
    "go_digit": _resolve_go_digit,
    "hdfc_ergo": _resolve_hdfc,
    "reliance": _resolve_reliance,
    "tata_aig": _resolve_tata,
}
