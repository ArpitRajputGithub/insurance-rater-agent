"""Pure-logic unit tests -- no OCR, no file I/O.

These lock the deterministic decisions that are easy to get subtly wrong:
premium-slab boundaries, segment banding, and the status roll-up precedence.
"""
from insurance_rater import grids
from insurance_rater.agent import _rollup
from insurance_rater.extract import _derive_type
from insurance_rater.models import ComponentRate, Status


def _cr(component, applicable, status):
    return ComponentRate(component, applicable, status)


def test_hdfc_slab_boundaries():
    assert grids._hdfc_slab(0) == "<10k"
    assert grids._hdfc_slab(9_999) == "<10k"
    assert grids._hdfc_slab(10_000) == ">10-50k"      # boundary is exclusive-low
    assert grids._hdfc_slab(50_000) == ">50k-1L"
    assert grids._hdfc_slab(100_000) == ">1L-2L"
    assert grids._hdfc_slab(200_000) == ">2L"
    assert grids._hdfc_slab(5_000_000) == ">2L"


def test_go_digit_segment_banding():
    assert grids._godigit_segment("petrol", 999) == "Petrol<1000"
    assert grids._godigit_segment("petrol", 1000) == "Petrol>1000"
    assert grids._godigit_segment("diesel", 1499) == "Diesel<1500"
    assert grids._godigit_segment("diesel", 1500) == "Diesel>1500"
    assert grids._godigit_segment("cng", 900) == "CNG"
    assert grids._godigit_segment(None, 1000) is None       # unknown fuel -> no guess


def test_derive_type_reads_title_not_the_type_of_cover_trap():
    # The Tata SATP fixture prints "Type of Cover: Package" on a Liability-Only
    # policy; the title must win so it stays satp, not flip to comprehensive.
    trap = ["Auto Secure - Liability Only Policy\n3. Type of Cover: Package"]
    assert _derive_type(trap) == "satp"
    assert _derive_type(["Digit Private Car Package Policy"]) == "comprehensive"
    assert _derive_type(["PRIVATE CAR COMPREHENSIVE POLICY"]) == "comprehensive"
    # No title and no content markers -> 'unknown', NOT a per-insurer default guess.
    assert _derive_type(["nothing type-bearing here"]) == "unknown"
    # SAOD title -> own-damage-only, not third-party.
    assert _derive_type(["Digit Private Car Stand-alone Own Damage Policy"]) == "saod"


def test_go_digit_saod_inverts_od_tp():
    # SAOD policy: OD applies but the TP-only card can't rate it (unsupported);
    # TP does not apply at all. The old resolver had this exactly backwards.
    from insurance_rater.grids.go_digit import resolve as _resolve_go_digit
    from insurance_rater.models import PolicyFacts
    od, tp, _ = _resolve_go_digit(PolicyFacts("Go Digit", "saod"), raters_dir="unused")
    assert od.applicable and od.status is Status.UNSUPPORTED
    assert not tp.applicable and tp.status is Status.RESOLVED


def test_llm_build_maps_fields_and_formats_rto():
    # The LLM path must produce the same PolicyFacts shape the resolvers read:
    # coerced ints, lowercased fuel, page citations, and the insurer-specific
    # rto_code form (Go Digit 'HR51' vs Reliance 'HR-51').
    from insurance_rater import llm
    raw = {"insurer": "go_digit", "policy_type": "saod", "fields": {
        "registration": {"value": "HR51CQ0040", "page": 1, "quote": "HR51CQ0040"},
        "cc": {"value": "2,184", "page": 1, "quote": "2184 CC"},
        "fuel": {"value": "Diesel", "page": 1},
    }}
    pf = llm._build(raw, "x.pdf")
    assert pf.insurer_key == "go_digit"
    assert pf.policy_type == "saod"
    assert pf.get("rto_code") == "HR51"
    assert pf.get("cc") == 2184
    assert pf.get("fuel") == "diesel"
    assert pf.facts["rto_code"].citation.locator == "page 1"
    pf2 = llm._build({"insurer": "reliance",
                      "fields": {"registration": {"value": "HR51CQ0040", "page": 1}}},
                     "x.pdf")
    assert pf2.get("rto_code") == "HR-51"


def test_reliance_cc_footnote_reads_both_numbers_from_the_card():
    # The threshold and % both live in the cell, and the card is replaced
    # monthly, so a future month changing either must be honoured without a
    # code change -- and unrelated T&C text must yield no rule.
    from insurance_rater.grids import _cc_footnote
    assert _cc_footnote("Payout for < 1000 CC, PO will be 5% lesser than the above Grid") == (1000, 5.0)
    assert _cc_footnote("Payout for < 1200 CC, PO will be 10% lesser than the above Grid") == (1200, 10.0)
    assert _cc_footnote("< 900 CC -> 2.5% less") == (900, 2.5)
    assert _cc_footnote("Standlone ZD policies 2.5% will be redused from the above Grid") is None
    assert _cc_footnote("") is None


def test_unknown_policy_type_refuses_to_resolve():
    # When the type can't be determined, the resolver must refuse (both
    # components unsupported) rather than fall into an else-branch that assumes
    # one -- the guard sits in resolve(), before any grid file is opened.
    from insurance_rater.grids import resolve
    from insurance_rater.models import PolicyFacts
    od, tp, _ = resolve(PolicyFacts("Go Digit", "unknown", "go_digit"), "go_digit", "unused")
    assert od.applicable and od.status is Status.UNSUPPORTED
    assert tp.applicable and tp.status is Status.UNSUPPORTED


def test_insurer_registry_is_the_single_source():
    # One registry drives detection, parsing and resolving: every entry must have
    # a parser and a resolver keyed the same way, and its fingerprint must detect
    # its own schedule. Adding/rebranding an insurer is then one registry edit.
    from insurance_rater import insurers
    from insurance_rater.extract import detect_insurer, _PARSERS
    from insurance_rater.grids import _RESOLVERS
    for ins in insurers.REGISTRY:
        assert ins.key in _PARSERS and ins.key in _RESOLVERS
        assert detect_insurer([f"{ins.name} General Insurance schedule"]) == ins.key
    # First-match-wins order: a competitor named in HDFC boilerplate must not win.
    assert detect_insurer(["HDFC ERGO ... Previous Policy of TATA AIG GENERAL "
                           "INSURANCE CO.LTD."]) == "hdfc_ergo"
    assert detect_insurer(["nothing identifying here"]) is None


def test_rollup_precedence():
    R, U, A = Status.RESOLVED, Status.UNSUPPORTED, Status.AMBIGUOUS
    # not-applicable OD (TP-only policy) is ignored in the roll-up
    assert _rollup(_cr("od", False, R), _cr("tp", True, R)) is R
    # a single unsupported applicable component pulls the whole run to unsupported
    assert _rollup(_cr("od", True, R), _cr("tp", True, U)) is U
    # ambiguous outranks unsupported
    assert _rollup(_cr("od", True, A), _cr("tp", True, U)) is A


def test_llm_call_retries_chain_after_quota_window(monkeypatch):
    # Every model 429s on the first pass; after one sleep the first model
    # answers. _call must fail over through the chain, wait once, and succeed.
    import io as _io
    import json as _json
    import urllib.error
    from insurance_rater import llm

    calls, sleeps = [], []
    ok = {"choices": [{"message": {"content": "{\"fields\": {}}"}}]}

    def fake_urlopen(req, timeout=0):
        calls.append(_json.loads(req.data)["model"])
        if len(calls) <= len(llm._MODELS):
            raise urllib.error.HTTPError(req.full_url, 429, "quota", {}, _io.BytesIO(b""))
        class R:
            def __enter__(self): return _io.StringIO(_json.dumps(ok))
            def __exit__(self, *a): pass
        return R()

    monkeypatch.setattr(llm.urllib.request, "urlopen", fake_urlopen)
    monkeypatch.setattr(llm.time, "sleep", sleeps.append)
    monkeypatch.setenv("GEMINI_API_KEY", "test")
    assert llm._call(["data:image/png;base64,x"]) == {"fields": {}}
    assert calls == list(llm._MODELS) + [llm._MODELS[0]]
    assert sleeps == [30]

    # A non-retryable error (401 bad key) must raise immediately, no retry.
    calls.clear()
    def fake_401(req, timeout=0):
        calls.append(1)
        raise urllib.error.HTTPError(req.full_url, 401, "unauthorized", {}, _io.BytesIO(b""))
    monkeypatch.setattr(llm.urllib.request, "urlopen", fake_401)
    try:
        llm._call(["data:image/png;base64,x"])
        assert False, "expected HTTPError"
    except urllib.error.HTTPError:
        pass
    assert calls == [1]


def test_onboard_validate_rejects_hallucinated_coordinates():
    # The proposer must not trust LLM-cited sheets/cells that don't exist.
    import openpyxl
    from insurance_rater.onboard import validate
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rates"
    ws["A1"], ws["H3"] = "header", "Payout for < 1000 CC, 5% lesser"
    good = {"sheets": [{"name": "Rates"}],
            "rate_sheets": [{"sheet": "Rates", "header_row": 1}],
            "rto_mappings": [{"sheet": "Rates"}],
            "footnotes": [{"sheet": "Rates", "cell": "H3"}]}
    assert validate(good, wb) == []
    bad = {"sheets": [{"name": "Ghost"}],
           "rate_sheets": [{"sheet": "Rates", "header_row": 999}],
           "rto_mappings": [{"sheet": "Nope"}],
           "footnotes": [{"sheet": "Rates", "cell": "ZZ99"},   # empty cell
                         {"sheet": "Rates", "cell": "not-a-cell"}]}
    issues = validate(bad, wb)
    assert len(issues) == 5, issues
